import requests
import pandas as pd
import os
import re
from openpyxl import load_workbook, Workbook
from datetime import datetime
import csv

# Carregar variáveis do arquivo .env
from dotenv import load_dotenv
load_dotenv()

# ========== CONFIGURAÇÕES ==========
JIRA_DOMAIN = 'https://vtex-dev.atlassian.net'
EMAIL = 'wilson.vilela@vtex.com'
API_TOKEN = os.getenv("JIRA_API_TOKEN")  # Token agora vem do .env
EXCEL_PATH = "jira_exportado_completo.xlsx"


TEAM_MAP = {
    "thago.oliveira@vtex.com": "SAP-Team",
    "milena.firmino@vtex.com": "SAP-Team",
    "wallace.costa@vtex.com": "SAP-Team",
    "raphaela.marinho@vtex.com": "Treasury-BRA",
    "wilson.vilela@vtex.com": "SAP-Team",
    "diego.almeida@vtex.com": "SAP-Team",
    "luan.freire@vtex.com": "SAP-Team",
    "fernanda.avila@vtex.com": "SAP-Team",
    "anrie.rodrigues@vtex.com": "SAP-Team",
    "mateus.vasconcelos@vtex.com": "SAP-Team"
}

# ========== PARSER ADF ==========
def parse_adf_to_text(adf):
    if not adf or not isinstance(adf, dict):
        return ""
    text_parts = []
    def extract_content(node):
        if node["type"] == "paragraph":
            line = ""
            for content in node.get("content", []):
                if content["type"] == "text":
                    line += content.get("text", "")
            text_parts.append(line)
        elif node["type"] == "bulletList":
            for item in node.get("content", []):
                for paragraph in item.get("content", []):
                    line = "- "
                    for content in paragraph.get("content", []):
                        if content["type"] == "text":
                            line += content.get("text", "")
                    text_parts.append(line)
        elif node.get("content"):
            for child in node["content"]:
                extract_content(child)
    for node in adf.get("content", []):
        extract_content(node)
    return "\n".join(text_parts)

# ========== SUPORTE ==========
def limpar_codigo_do_resumo(texto):
    if not texto:
        return ""
    return re.sub(r"^\(?\d{5}\)?\s*-?\s*", "", texto)

def extrair_codigo(texto):
    if not texto:
        return ""
    match = re.search(r"\(?(\d{5})\)?", texto)
    return match.group(1) if match else ""

# ========== COLETA DO JIRA ==========
auth = (EMAIL, API_TOKEN)
headers = {"Accept": "application/json"}

JQL = "project = SAPS ORDER BY created DESC"
start_at = 0
max_results = 100
issues_data = []

print("Iniciando exportação...")

while True:
    url = f"{JIRA_DOMAIN}/rest/api/3/search"
    params = {
        "jql": JQL,
        "startAt": start_at,
        "maxResults": max_results,
        "fields": "key,summary,created,updated,description,priority,reporter,assignee,status,customfield_10016,customfield_10704"
    }
    response = requests.get(url, headers=headers, params=params, auth=auth)
    if response.status_code != 200:
        print("Erro:", response.status_code, response.text)
        break
    data = response.json()
    issues = data.get("issues", [])
    for issue in issues:
        fields = issue.get("fields", {})
        reporter = fields.get("reporter")
        assignee = fields.get("assignee")
        descricao = parse_adf_to_text(fields.get("description"))
        resumo = limpar_codigo_do_resumo(fields.get("summary"))
        email = reporter.get("emailAddress") if reporter else ""
        custom_team = fields.get("customfield_10704")
        time = (custom_team.get("value") if isinstance(custom_team, dict) else None) or TEAM_MAP.get(email, "")
        issues_data.append({
            "Chave": issue.get("key"),
            "Criado": fields.get("created"),
            "Prioridade": fields.get("priority", {}).get("name") if fields.get("priority") else None,
            "Resumo": resumo,
            "Relator": reporter.get("displayName") if reporter else None,
            "Responsável": assignee.get("displayName") if assignee else None,
            "Time": time,
            "Status": fields.get("status", {}).get("name") if fields.get("status") else None,
            "Atualizado(a)": fields.get("updated"),
            "Time to resolution": fields.get("customfield_10016"),
            "Descrição": descricao,
            "CONTAGEM": "1",
            "CÓDIGO": extrair_codigo(fields.get("summary", ""))
        })
    if data.get("total", 0) <= start_at + max_results:
        break
    start_at += max_results

df = pd.DataFrame(issues_data)

# ========== RECARREGAR KEYWORDS SEM ATUALIZAR ==========
if os.path.exists(EXCEL_PATH):
    antigo = pd.read_excel(EXCEL_PATH, sheet_name=None)
    antiga_base = antigo.get("JIRA")
    if antiga_base is not None and "Chave" in antiga_base.columns and "KEYWORD" in antiga_base.columns:
        df = df.merge(antiga_base[["Chave", "KEYWORD"]], on="Chave", how="left")
    else:
        df["KEYWORD"] = ""
else:
    df["KEYWORD"] = ""

# ========== ATUALIZAÇÃO COMPLETA DA ABA JIRA OU CRIAÇÃO DO ARQUIVO ==========
if os.path.exists(EXCEL_PATH):
    wb = load_workbook(EXCEL_PATH)
    if "JIRA" in wb.sheetnames:
        ws = wb["JIRA"]
        cabecalho = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
        colunas_excel = list(cabecalho.keys())

        for row_idx, (_, linha) in enumerate(df.iterrows(), start=2):
            for col_idx, col_nome in enumerate(colunas_excel, start=1):
                valor = linha.get(col_nome, "")
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = valor

        for row in ws.iter_rows(min_row=len(df) + 2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
    else:
        print("A aba 'JIRA' não existe. Crie o arquivo Excel com a estrutura inicial.")
        exit(1)
else:
    print("Arquivo Excel não encontrado. Criando novo arquivo...")
    wb = Workbook()
    ws = wb.active
    ws.title = "JIRA"
    colunas_jira = list(df.columns)
    for idx, col_nome in enumerate(colunas_jira, start=1):
        ws.cell(row=1, column=idx).value = col_nome
    for row_idx, row in enumerate(df[colunas_jira].values.tolist(), start=2):
        for col_idx, valor in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = valor
    ws2 = wb.create_sheet("Descricao_Segmentada")
    df_segmentado = df[["Chave", "CÓDIGO", "Descrição"]].copy()
    descricoes = df_segmentado["Descrição"].fillna("").astype(str).str.split("#")
    for i in range(1, 10):
        df_segmentado[f"Descricao{i+1}"] = descricoes.apply(lambda x: x[i] if len(x) > i else "")
    df_segmentado = df_segmentado.drop(columns=["Descrição"])
    for col_idx, col_nome in enumerate(df_segmentado.columns, start=1):
        ws2.cell(row=1, column=col_idx).value = col_nome
    for row_idx, row in enumerate(df_segmentado.values, start=2):
        for col_idx, valor in enumerate(row, start=1):
            ws2.cell(row=row_idx, column=col_idx).value = valor
    wb.save(EXCEL_PATH)
    print("Novo arquivo Excel criado com abas 'JIRA' e 'Descricao_Segmentada'. Continuação da execução...")

# ========== ATUALIZAÇÃO DA ABA DESCRICAO_SEGMENTADA ==========
if "Descricao_Segmentada" in wb.sheetnames:
    ws2 = wb["Descricao_Segmentada"]
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=ws2.max_column):
        for cell in row:
            cell.value = None
else:
    ws2 = wb.create_sheet("Descricao_Segmentada")

df_segmentado = df[["Chave", "CÓDIGO", "Descrição"]].copy()
descricoes = df_segmentado["Descrição"].fillna("").astype(str).str.split("#")
for i in range(1, 10):
    df_segmentado[f"Descricao{i+1}"] = descricoes.apply(lambda x: x[i] if len(x) > i else "")
df_segmentado = df_segmentado.drop(columns=["Descrição"])

for col_idx, col_nome in enumerate(df_segmentado.columns, start=1):
    ws2.cell(row=1, column=col_idx).value = col_nome
for row_idx, row in enumerate(df_segmentado.values, start=2):
    for col_idx, valor in enumerate(row, start=1):
        ws2.cell(row=row_idx, column=col_idx).value = valor

wb.save(EXCEL_PATH)
print("Exportação concluída com sucesso e com formatação preservada.")

# ========== CÁLCULO DE MÉTRICAS ==========
# ========== CÁLCULO DE MÉTRICAS ==========
novos = 0
alterados_desc = 0
alterados_status = 0
alterados_responsavel = 0
inalterados = 0

if 'antiga_base' in locals() and antiga_base is not None and "Chave" in antiga_base.columns:
    antiga_base = antiga_base.set_index("Chave")
    nova_base = df.set_index("Chave")
    for chave in nova_base.index:
        if chave not in antiga_base.index:
            novos += 1
        else:
            alterou_desc = nova_base.at[chave, "Descrição"] != antiga_base.at[chave, "Descrição"]
            alterou_status = nova_base.at[chave, "Status"] != antiga_base.at[chave, "Status"]
            alterou_resp = nova_base.at[chave, "Responsável"] != antiga_base.at[chave, "Responsável"]
            if alterou_desc:
                alterados_desc += 1
            if alterou_status:
                alterados_status += 1
            if alterou_resp:
                alterados_responsavel += 1
            if not alterou_desc and not alterou_status and not alterou_resp:
                inalterados += 1
else:
    novos = len(df)

# ========== LOG EM CSV ==========
log_csv_path = "log_execucao.csv"
arquivo_novo = not os.path.exists(log_csv_path)

with open(log_csv_path, "a", newline="", encoding="utf-8") as log_file:
    writer = csv.writer(log_file)
    if arquivo_novo:
        writer.writerow(["Data", "Novos", "Descricao Alterada", "Status Alterado", "Responsavel Alterado", "Inalterados"])
    writer.writerow([
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        novos,
        alterados_desc,
        alterados_status,
        alterados_responsavel,
        inalterados
    ])
    print("Log CSV de execução registrado.")

# ========== LOG DETALHADO EM TXT ==========
log_txt_path = "log_detalhado.txt"
agora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

with open(log_txt_path, "a", encoding="utf-8") as log_txt:
    log_txt.write(f"[{agora}] Início da exportação\n")
    log_txt.write(f"[{agora}] Tickets novos: {novos}\n")
    log_txt.write(f"[{agora}] Tickets com descrição alterada: {alterados_desc}\n")
    log_txt.write(f"[{agora}] Tickets com status alterado: {alterados_status}\n")
    log_txt.write(f"[{agora}] Tickets com responsável alterado: {alterados_responsavel}\n")
    log_txt.write(f"[{agora}] Tickets inalterados: {inalterados}\n")
    log_txt.write(f"[{agora}] Fim da execução\n\n")

print("\nResumo final da execução:")
print(f" - Tickets novos: {novos}")
print(f" - Descrição alterada: {alterados_desc}")
print(f" - Status alterado: {alterados_status}")
print(f" - Responsável alterado: {alterados_responsavel}")
print(f" - Inalterados: {inalterados}")
print("Log detalhado em TXT registrado.")
